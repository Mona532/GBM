from __future__ import annotations

import csv
import gzip
import sys
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_MAP = {
    "tissue_name": "top_level",
    "mixed_tissue_name": "top_level_mixed_name",
    "cell_name_standardized": "cell_name_standardized",
    "cell_name_cl_id": "cell_name_cl",
    "disease_type_do": "disease_type_do",
    "disease_type": "disease_type",
    "marker_polarity": "marker_polarity",
    "gene_symbol": "gene_symbol",
    "species": "species",
}


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: prepare_scmarkeragent_db_from_excel.py <input_xlsx> <output_csv_gz>")
        return 2

    input_xlsx = Path(sys.argv[1])
    output_csv_gz = Path(sys.argv[2])
    output_csv_gz.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(x) if x is not None else "" for x in header]
    index = {name: i for i, name in enumerate(header)}

    missing = [src for src in REQUIRED_MAP if src not in index]
    if missing:
      print(f"missing columns: {missing}")
      return 1

    out_cols = list(REQUIRED_MAP.values())
    row_count = 0

    with gzip.open(output_csv_gz, "wt", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(out_cols)
        for row in ws.iter_rows(min_row=2, values_only=True):
            out_row = []
            for src in REQUIRED_MAP:
                value = row[index[src]]
                if value is None:
                    out_row.append("")
                else:
                    out_row.append(str(value))
            writer.writerow(out_row)
            row_count += 1
            if row_count % 100000 == 0:
                print(f"rows={row_count}")

    print(f"done rows={row_count} output={output_csv_gz}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
